import openpyxl

#주소 변수에 저장
home="/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/test.xlsx/test.xlsx"
# 엑셀 불러오기
wb= openpyxl.load_workbook(home)

# 엑셀 시트 선택
ws=wb['sheet1']

#데이터 수정하기
ws['A3']= "X"
ws['B3']= "XXXXXXX"
ws['C3']="테스트 숴정"
ws['D3']= "X"
ws['E3']= "XXX"
ws['F3']= ""
ws['G3']= "X"

#엑셀 저장
wb.save(home)