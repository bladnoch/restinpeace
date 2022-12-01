from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo


#from datetime improt datetime

#def time():

def myFunc(): #DK 테스트용 재출력 버튼에 연결되어 있음
    messagebox.showinfo("버튼 클릭","ID : "+ID.get()+"\n고인명 : "+고인명.get()+"\n 상주명 : "+상주명.get())



def show_excel(): #DK 리스트에 한줄로 insert
    wb = openpyxl.load_workbook(home, data_only=True) #값만 받기
    ws = wb['Sheet1']

    one_line = "" #DK 이 변수에 한줄 저장
    for k in range(1, 8):
        if (str(ws.cell(row=ct, column=k).value)=="None"): #DK G의 함수를 None -> 0으로 받기
            one_line+="0"
        else:
           one_line += str(ws.cell(row=ct, column=k).value)+'  '
    return one_line


#DK og 목록을 참조한 2차원 리스트
#DK 2차원 리스트에 저장
def in_list():
    wb = openpyxl.load_workbook(home)
    ws = wb['Sheet1']

    for i in range(1, 41):
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        col.append(row);
        row = []

#DK add row(not yet)
    #DK 저장된 2차원 리스트를 볼 수 있음
    # for i in range(40):
    #     for j in range(7):
    #         print(col[i][j], end="/t")
    #     print("")

    # def gotable():
    #     # wb = openpyxl.load_workbook(home)
    #     # ws = wb['Sheet1']
    #
    #     wb = Workbook()
    #     ws = wb.active
    #
    #     data = [
    #         ['Apples', 10000, 5000, 8000, 6000],
    #         ['Pears',   2000, 3000, 4000, 5000],
    #         ['Bananas', 6000, 6000, 6500, 6000],
    #         ['Oranges',  500,  300,  200,  700],
    #     ]
    #     data=col
    #
    #
    #     # add column headings. NB. these must be strings
    #     ws.append(["Fruit", "2011", "2012", "2013", "2014"])
    #     for row in data:
    #         ws.append(row)
    #
    #     tab = Table(displayName="Table1", ref="A1:E5")
    #
    #     # Add a default style with striped rows and banded columns
    #     style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
    #                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    #     tab.tableStyleInfo = style
    #
    #     '''
    #     Table must be added using ws.add_table() method to avoid duplicate names.
    #     Using this method ensures table name is unque through out defined names and all other table name.
    #     '''
    #     ws.add_table(tab)
    #     wb.save("copyog.xlsx")

def close():
    win.quit()
    win.destroy()

#########################   global variable   ##########################

home = "/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/test.xlsx" #기본 물품 엑셀 위치 저장

#DK 초기 excel 파일용 전역변수
row=[] #2차원 리스트에 값 저장할 때 사용
col=[] #2차원 리스트에 값 저장할 때 사용
ct=1 #row 카운트 할때 사용
i=0 #리스트 인서트할 때 for에 사용
one_line="" #한줄로 받을 때 사용

#Tkinter 윈도우 화면
win = Tk() # 창 생성
win.geometry("1000x720") # 창의 크기
win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
win.option_add("*Font", "맑은고딕 11") # 전체 폰트
#win.resizable(False, False) #윈도우 사이즈 조절 불가

#########################   excel   ##########################

#########################    menu   ##########################

menu = Menu(win)

menu_1 = Menu(menu, tearoff = 0)
menu_1.add_command(label = "로그인")
menu_1.add_command(label = "인쇄")
menu_1.add_command(label = "장부")
menu_1.add_separator()
menu_1.add_command(label = "종료", command = close)
menu.add_cascade(label = "메뉴", menu = menu_1)

win.config(menu = menu)

#########################   config  ##########################

#레이블 정의
ID_lab = Label(win)
ID_lab.config(text = "ID", width=10, relief="solid")
고인명_lab = Label(win)
고인명_lab.config(text = "고인명",width=10, relief="solid")
상주명_lab = Label(win)
상주명_lab.config(text = "상주명",width=10, relief="solid")
빈소_lab = Label(win)
빈소_lab.config(text = "빈소", width=10, relief="solid")
빈소기간_lab = Label(win)
빈소기간_lab.config(text = "빈소기간", width=10, relief="solid")
안치기간_lab = Label(win)
안치기간_lab.config(text = "안치기간", width=10, relief="solid")
물결1_lab = Label(win)
물결1_lab.config(text = "~", width=10)
물결2_lab = Label(win)
물결2_lab.config(text = "~", width=10)
###
수납금액_lab = Label(win)
수납금액_lab.config(text = "수납금액", width=10, relief="solid")
받은금액_lab = Label(win)
받은금액_lab.config(text = "받은금액", width=10, relief="solid")
거스름돈_lab = Label(win)
거스름돈_lab.config(text = "거스름돈", width=10, relief="solid")

#엔트리 정의
ID = Entry(win)
ID.config(width=10,relief="solid",borderwidth=2)
고인명 = Entry(win)
고인명.config(width=10,relief="solid",borderwidth=2)
상주명 = Entry(win)
상주명.config(width=10,relief="solid",borderwidth=2)
빈소 = Entry(win)
빈소.config(width=60,relief="solid",borderwidth=2)
빈소기간1 = Entry(win)
빈소기간1.config(width=20,relief="solid",borderwidth=2)
안치기간1 = Entry(win)
안치기간1.config(width=20,relief="solid",borderwidth=2)
빈소기간2 = Entry(win)
빈소기간2.config(width=20,relief="solid",borderwidth=2)
안치기간2 = Entry(win)
안치기간2.config(width=20,relief="solid",borderwidth=2)
수납금액 = Entry(win)
수납금액.config(width=20,relief="solid",borderwidth=2)
받은금액 = Entry(win)
받은금액.config(width=20,relief="solid",borderwidth=2)
거스름돈 = Entry(win)
거스름돈.config(width=20,relief="solid",borderwidth=2)

#버튼 정의
재출력 = Button(win, text = "재출력", command=myFunc) #DK command로 버튼 클릭시 def myDunc() 실행
재출력.config(width=10,height=2)
#btn.config(command=ID_a)
현금수납 = Button(win, text = "현금수납")
현금수납.config(width=10,height=3)
닫기 = Button(win, text = "닫기")
닫기.config(width=10,height=3,command =close)
식당판매 = Button(win, text = "식당판매")
식당판매.config(width=10,height=3)
매점판매 = Button(win, text = "매점판매")
매점판매.config(width=10,height=3)
Set = Button(win, text = "기본 Set")
Set.config(width=10,height=3)

in_list

리스트 = Listbox(win, selectmode = 'extended',width = 122, height = 30,)
리스트.yview()

# 리스트.gotable()

#DK 엑셀 저장된 2차원 리스트 불러오기
for i in range(40):
    리스트.insert(i,show_excel())
    ct+=1

#########################   place  ##########################

#레이블 위치
ID_lab.place(x=10,y=10)
고인명_lab.place(x=210,y=10)
상주명_lab.place(x=410,y=10)
빈소_lab.place(x=10,y=50)
빈소기간_lab.place(x=10,y=90)
안치기간_lab.place(x=10,y=130)
물결1_lab.place(x=250,y=90)
물결2_lab.place(x=250,y=130)
###
수납금액_lab.place(x=620,y=10)
받은금액_lab.place(x=620,y=60)
거스름돈_lab.place(x=620,y=110)

#엔트리 위치
ID.place(x=110,y=10)
고인명.place(x=310,y=10)
상주명.place(x=510,y=10)
빈소.place(x=110,y=50)
빈소기간1.place(x=110,y=90)
안치기간1.place(x=110,y=130)
빈소기간2.place(x=310,y=90)
안치기간2.place(x=310,y=130)

수납금액.place(x=720,y=10)
받은금액.place(x=720,y=60)
거스름돈.place(x=720,y=110)

#버튼 위치
재출력.place(x= 500, y=100)
현금수납.place(x=900, y=10)
닫기.place(x=900, y=70)
식당판매.place(x=700, y=150)
매점판매.place(x=800, y=150)
Set.place(x=900, y=150)

#리스트 위치
리스트.place(x=10, y=210)

win.mainloop() # 창 실행







############################################ 원본 파일 ############################################

# from tkinter import * # tkinter의 모든 함수 가져오기
# import openpyxl
# #from datetime improt datetime
#
# #함수 정의 부분
# #def time():
#
# #def ID_a():
#
# def close():
#     win.quit()
#     win.destroy()
#
# #Tkinter 윈도우 화면
# win = Tk() # 창 생성
# win.geometry("1000x720") # 창의 크기
# win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
# win.option_add("*Font", "맑은고딕 11") # 전체 폰트
# #win.resizable(False, False) #윈도우 사이즈 조절 불가
#
# #########################   excel   ##########################
#
# #########################    menu   ##########################
#
# menu = Menu(win)
#
# menu_1 = Menu(menu, tearoff = 0)
# menu_1.add_command(label = "로그인")
# menu_1.add_command(label = "인쇄")
# menu_1.add_command(label = "장부")
# menu_1.add_separator()
# menu_1.add_command(label = "종료", command = close)
# menu.add_cascade(label = "메뉴", menu = menu_1)
#
# win.config(menu = menu)
#
# #########################   config  ##########################
#
# #레이블 정의
# ID_lab = Label(win)
# ID_lab.config(text = "ID",width=10, relief="solid")
# 고인명_lab = Label(win)
# 고인명_lab.config(text = "고인명",width=10, relief="solid")
# 상주명_lab = Label(win)
# 상주명_lab.config(text = "상주명",width=10, relief="solid")
# 빈소_lab = Label(win)
# 빈소_lab.config(text = "빈소", width=10, relief="solid")
# 빈소기간_lab = Label(win)
# 빈소기간_lab.config(text = "빈소기간", width=10, relief="solid")
# 안치기간_lab = Label(win)
# 안치기간_lab.config(text = "안치기간", width=10, relief="solid")
# 물결1_lab = Label(win)
# 물결1_lab.config(text = "~", width=10)
# 물결2_lab = Label(win)
# 물결2_lab.config(text = "~", width=10)
# ###
# 수납금액_lab = Label(win)
# 수납금액_lab.config(text = "수납금액", width=10, relief="solid")
# 받은금액_lab = Label(win)
# 받은금액_lab.config(text = "받은금액", width=10, relief="solid")
# 거스름돈_lab = Label(win)
# 거스름돈_lab.config(text = "거스름돈", width=10, relief="solid")
#
# #엔트리 정의
# ID = Entry(win)
# ID.config(width=10,relief="solid",borderwidth=2)
# 고인명 = Entry(win)
# 고인명.config(width=10,relief="solid",borderwidth=2)
# 상주명 = Entry(win)
# 상주명.config(width=10,relief="solid",borderwidth=2)
# 빈소 = Entry(win)
# 빈소.config(width=60,relief="solid",borderwidth=2)
# 빈소기간1 = Entry(win)
# 빈소기간1.config(width=20,relief="solid",borderwidth=2)
# 안치기간1 = Entry(win)
# 안치기간1.config(width=20,relief="solid",borderwidth=2)
# 빈소기간2 = Entry(win)
# 빈소기간2.config(width=20,relief="solid",borderwidth=2)
# 안치기간2 = Entry(win)
# 안치기간2.config(width=20,relief="solid",borderwidth=2)
# 수납금액 = Entry(win)
# 수납금액.config(width=20,relief="solid",borderwidth=2)
# 받은금액 = Entry(win)
# 받은금액.config(width=20,relief="solid",borderwidth=2)
# 거스름돈 = Entry(win)
# 거스름돈.config(width=20,relief="solid",borderwidth=2)
#
# #버튼 정의
# 재출력 = Button(win, text = "재출력")
# 재출력.config(width=10,height=2)
# #btn.config(command=ID_a)
# 현금수납 = Button(win, text = "현금수납")
# 현금수납.config(width=10,height=3)
# 닫기 = Button(win, text = "닫기")
# 닫기.config(width=10,height=3,command =close)
# 식당판매 = Button(win, text = "식당판매")
# 식당판매.config(width=10,height=3)
# 매점판매 = Button(win, text = "매점판매")
# 매점판매.config(width=10,height=3)
# Set = Button(win, text = "기본 Set")
# Set.config(width=10,height=3)
#
# 리스트 = Listbox(win, selectmode = 'extended',width = 122, height = 30,)
# 리스트.yview()
# 리스트.insert(0, "1번")
# 리스트.insert(1, "2번") #반복문으로 딕셔너리, 튜플, 리스트 사용 가
#
# #########################   place  ##########################
#
# #레이블 위치
# ID_lab.place(x=10,y=10)
# 고인명_lab.place(x=210,y=10)
# 상주명_lab.place(x=410,y=10)
# 빈소_lab.place(x=10,y=50)
# 빈소기간_lab.place(x=10,y=90)
# 안치기간_lab.place(x=10,y=130)
# 물결1_lab.place(x=250,y=90)
# 물결2_lab.place(x=250,y=130)
# ###
# 수납금액_lab.place(x=620,y=10)
# 받은금액_lab.place(x=620,y=60)
# 거스름돈_lab.place(x=620,y=110)
#
# #엔트리 위치
# ID.place(x=110,y=10)
# 고인명.place(x=310,y=10)
# 상주명.place(x=510,y=10)
# 빈소.place(x=110,y=50)
# 빈소기간1.place(x=110,y=90)
# 안치기간1.place(x=110,y=130)
# 빈소기간2.place(x=310,y=90)
# 안치기간2.place(x=310,y=130)
#
# 수납금액.place(x=720,y=10)
# 받은금액.place(x=720,y=60)
# 거스름돈.place(x=720,y=110)
#
# #버튼 위치
# 재출력.place(x= 500, y=100)
# 현금수납.place(x=900, y=10)
# 닫기.place(x=900, y=70)
# 식당판매.place(x=700, y=150)
# 매점판매.place(x=800, y=150)
# Set.place(x=900, y=150)
#
# #리스트 위치
# 리스트.place(x=10, y=210)
#
# win.mainloop() # 창 실행
